from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from time import sleep
from datetime import datetime
import openpyxl
import os
import schedule

#criar a planilha

def criar_planilha(nome_planilha,nome_sheet,cabecalho):
    workbook = openpyxl.Workbook()
    del workbook['Sheet']
    workbook.create_sheet(nome_sheet)
    sheet_pesquisa = workbook[nome_sheet]
    sheet_pesquisa.append(cabecalho)
    workbook.save(nome_planilha)

#inserir dados na planilha, caso exista,se não existir cria antes de inserir
    
def inserir_dados(nome_planilha,nome_sheet,dados):
    
    if not os.path.exists(nome_planilha):
        criar_planilha(nome_planilha,nome_sheet,['Produto','Data','Valor','Link'])
        
    workbook = openpyxl.load_workbook(nome_planilha)
    sheet_pesquisa = workbook[nome_sheet]
    sheet_pesquisa.append(dados)
    workbook.save(nome_planilha)
        
#inicia o driver   
def iniciar_driver():
    chrome_options = Options()
    arguments = ['--lang=pt-BR', '--window-size=800,700','--incognito' ]
    for argument in arguments:
        chrome_options.add_argument(argument)

    chrome_options.add_experimental_option('prefs', {
        'download.prompt_for_download': False,
        'profile.default_content_setting_values.notifications': 2,
        'profile.default_content_setting_values.automatic_downloads': 1,

    })
    driver = webdriver.Chrome(options=chrome_options)
    
    return driver

#obtém apenas o preço do produto
def obter_preco(driver):
    container_preco = driver.find_element(By.CLASS_NAME,'ui-pdp-price__main-container')
    preco = container_preco.find_element(
        By.CLASS_NAME,'ui-pdp-price__second-line').find_element(
            By.CLASS_NAME,'andes-money-amount.ui-pdp-price__part.andes-money-amount--cents-superscript.andes-money-amount--compact'
            ).text.replace('\n','').replace('R$','')
    return preco

#formata o preco de acordo com o valor dos centavos
def formatar_preco(preco):
    lista_inteiro_centavos = preco.split(',')
    inteiro = lista_inteiro_centavos[0].replace('.','')
    centavos = lista_inteiro_centavos[1]
    if centavos == '00':
        return inteiro
    else:
        return inteiro + '.' + centavos
    

def monitorar_preco():

    try:
        
        link_produto = 'https://www.mercadolivre.com.br/macbook-air-m1-2020-133-cinza-espacial-8gb-de-ram-256gb-ssd-apple-m-distribuidor-autorizado/p/MLB1018721006?pdp_filters=item_id:MLB2205623698#is_advertising=true&searchVariation=MLB1018721006&position=1&search_layout=grid&type=pad&tracking_id=e864f426-ac37-4e1f-b9ac-e436c026d4ba&is_advertising=true&ad_domain=VQCATCORE_LST&ad_position=1&ad_click_id=YjU4Y2Q5MjItOGE1OC00ODg4LWE5N2EtOWRkNWYzZTg3ZDU3'

        driver = iniciar_driver()

        driver.get(link_produto)

        sleep(5)

        produto = driver.find_element(By.TAG_NAME,'h1').text

        preco = obter_preco(driver)

        preco_formatado = formatar_preco(preco)

        data = datetime.now().strftime("%d-%m-%Y %H:%M")

        dados =[produto,data,preco_formatado,link_produto]

        inserir_dados('preco_produto.xlsx','pesquisa',dados)

        print(f'''

{100*"-"}

Pesquisa executada com sucesso às {data}

{100*"-"}

''')
        
    except Exception as e:
        
        data = datetime.now().strftime("%d-%m-%Y %H:%M")

        print(f'''

{100*"-"}

{e}

Erro ocorrido em: {data}

Próximo monitoramento daqui a 30 minutos

{100*"-"}

            ''')
    finally:
        
        driver.quit()
        
        


monitorar_preco()

schedule.every(30).minutes.do(monitorar_preco)

while True:
    
    schedule.run_pending()
    
    sleep(20)
